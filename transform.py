import pandas as pd
from extract import migracao_clientes, migracao_processos_1, migracao_processos_2
from unidecode import unidecode
import re
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import numpy as np


pd.options.display.float_format = '{:.0f}'.format

'''Função para ajustar o gênero em um texto. Altera 'bras' para 'brasileira' ou 'brasileiro' dependendo da terminação.'''
def ajustar_genero(texto):
    if pd.isna(texto):
        return texto
    if texto.endswith('a'):
        return texto.replace('bras', 'brasileira')
    elif texto.endswith('o'):
        return texto.replace('bras', 'brasileiro')
    else:
        return texto.replace('bras', 'brasileiro')


'''Função para corrigir caracteres especiais em um texto usando a biblioteca unidecode.'''
def corrigir_caracteres(texto):
    if pd.isnull(texto):
        return texto
    return unidecode(texto)


'''Função para formatar o número do processo. Preenche com zeros à esquerda e aplica um formato específico.'''
def formatar_numero_processo(valor):
    if pd.isnull(valor):
        return valor
    valor_str = str(valor)
    if valor_str.isdigit():
        valor_str = valor_str.zfill(20)
        return f"{valor_str[:7]}-{valor_str[7:9]}.{valor_str[9:13]}.{valor_str[13]}.{valor_str[14:16]}.{valor_str[16:]}"
    else:
        try:
            return valor_str.encode('mac_roman').decode('utf-8')
        except UnicodeEncodeError:
            return valor_str


regex_mascara = r'^\d{7}-\d{2}\.\d{4}\.\d{1}\.\d{2}\.\d{4}$'


def verificar_processo(valor):
    if pd.isnull(valor):
        return valor
    if valor == 'Aguardando numeração' or re.match(regex_mascara, valor):
        return valor
    else:
        return np.nan

'''Função principal para transformar o DataFrame de clientes. Aplica transformações de formatação e salva em um arquivo Excel.'''
def transformacao_clientes():
    _, df_modelo_clientes = migracao_clientes()
    df_modelo_clientes = df_modelo_clientes.dropna(subset=['NOME'])
    df_modelo_clientes = df_modelo_clientes.drop_duplicates(subset='NOME', keep='first')
    if 'NACIONALIDADE' in df_modelo_clientes.columns:
        df_modelo_clientes['NACIONALIDADE'] = df_modelo_clientes['NACIONALIDADE'].apply(ajustar_genero)
    df_modelo_clientes['DATA DE NASCIMENTO'] = pd.to_datetime(df_modelo_clientes['DATA DE NASCIMENTO'], format='%d/%m/%Y %H:%M')
    df_modelo_clientes['DATA DE NASCIMENTO'] = df_modelo_clientes['DATA DE NASCIMENTO'].dt.strftime('%d/%m/%Y')
    df_modelo_clientes['PROFISSÃO'] = df_modelo_clientes['PROFISSÃO'].apply(lambda x: x.encode('mac_roman').decode('utf-8') if pd.notnull(x) else x)
    df_modelo_clientes['PROFISSÃO'] = df_modelo_clientes['PROFISSÃO'].apply(corrigir_caracteres)
    df_modelo_clientes['ESTADO'] = df_modelo_clientes['ESTADO'].apply(lambda x: x.upper() if pd.notnull(x) else x)
    df_modelo_clientes['CIDADE'] = df_modelo_clientes['CIDADE'].replace('Floripa', 'Florianopolis')
    df_modelo_clientes['CIDADE'] = df_modelo_clientes['CIDADE'].apply(lambda x: x.encode('mac_roman').decode('utf-8') if pd.notnull(x) else x)
    df_modelo_clientes['BAIRRO'] = df_modelo_clientes['BAIRRO'].apply(lambda x: x.encode('mac_roman').decode('utf-8') if pd.notnull(x) else x)
    df_modelo_clientes['CEP'] = df_modelo_clientes['CEP'].apply(
        lambda x: f'{re.sub("[^0-9]", "", str(x))[:5].zfill(5)}-{re.sub("[^0-9]", "", str(x))[5:].zfill(3)}' 
        if pd.notnull(x) and len(re.sub("[^0-9]", "", str(x))) >= 8 else x
    )

    df_modelo_clientes['PIS PASEP'] = df_modelo_clientes['PIS PASEP'].apply(
        lambda x: f'{re.sub("[^0-9]", "", str(x))[:3].zfill(3)}.{re.sub("[^0-9]", "", str(x))[3:7].zfill(4)}.{re.sub("[^0-9]", "", str(x))[7:10].zfill(3)}-{re.sub("[^0-9]", "", str(x))[10:11].zfill(1)}'
        if pd.notnull(x) and len(re.sub("[^0-9]", "", str(x))) >= 11 else x
    )

    output_dir = 'data/output'
    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df_modelo_clientes, index=False, header=True):
        ws.append(row)
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(name='Arial', bold=True, size=10)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
    file_path = os.path.join(output_dir, 'CLIENTES.xlsx')
    wb.save(file_path)


'''Função principal para transformar o DataFrame de processos. Aplica transformações de formatação e salva em um arquivo Excel.'''
def transformacao_processos():
    df_modelo_processo, df_processos_migracao, df_grupo_processo, df_comarca = migracao_processos_1()
    df_fase_processo, df_grupo_processo = migracao_processos_2()
    df_clientes_migracao, _ = migracao_clientes()

    df_modelo_processo['NOME DO CLIENTE'] = df_processos_migracao['cod_cliente'].map(
        dict(zip(df_clientes_migracao['codigo'], df_clientes_migracao['razao_social']))
    )
    
    df_modelo_processo['GRUPO DE AÇÃO'] = df_processos_migracao['grupo_processo'].map(
        dict(zip(df_grupo_processo['codigo'], df_grupo_processo['descricao']))
    )
    df_modelo_processo['GRUPO DE AÇÃO'] = df_modelo_processo['GRUPO DE AÇÃO'].apply(lambda x: x.encode('mac_roman').decode('utf-8') if pd.notnull(x) else x)

    df_modelo_processo['NÚMERO DO PROCESSO'] = df_processos_migracao['numero_processo'].astype(str)
    df_modelo_processo['NÚMERO DO PROCESSO'] = df_modelo_processo['NÚMERO DO PROCESSO'].apply(formatar_numero_processo).replace('1,23451E+17', 'Aguardando numeração')
    df_modelo_processo['NÚMERO DO PROCESSO'] = df_modelo_processo['NÚMERO DO PROCESSO'].apply(verificar_processo)

    df_modelo_processo['COMARCA'] = df_processos_migracao['codcomarca'].map(
        dict(zip(df_comarca['codigo'], df_comarca['descricao']))
    )
    df_modelo_processo['DATA CADASTRO'] = pd.to_datetime(df_processos_migracao['data_contratacao'], format='%d/%m/%Y %H:%M', errors='coerce')
    df_modelo_processo['DATA CADASTRO'] = df_modelo_processo['DATA CADASTRO'].dt.strftime('%d/%m/%Y')

    df_modelo_processo['VALOR HONORÁRIOS'] = df_processos_migracao['valor_causa']

    df_modelo_processo['DATA FECHAMENTO'] = pd.to_datetime(df_processos_migracao['data_ultima_visualizacao'], format='%d/%m/%Y %H:%M', errors='coerce')
    df_modelo_processo['DATA FECHAMENTO'] = df_modelo_processo['DATA FECHAMENTO'].dt.strftime('%d/%m/%Y')

    df_modelo_processo['FASE PROCESSUAL'] = df_processos_migracao['codigo_fase'].map(
        dict(zip(df_fase_processo['codigo'], df_fase_processo['fase']))
    )
    df_modelo_processo['FASE PROCESSUAL'] = df_modelo_processo['FASE PROCESSUAL'].apply(lambda x: x.encode('mac_roman').decode('utf-8') if pd.notnull(x) else x)

    df_modelo_processo['GRUPO DE AÇÃO'] = df_processos_migracao['grupo_processo'].map(
        dict(zip(df_grupo_processo['codigo'], df_grupo_processo['descricao']))
    )
    df_modelo_processo['GRUPO DE AÇÃO'] = df_modelo_processo['GRUPO DE AÇÃO'].apply(lambda x: x.encode('mac_roman').decode('utf-8') if pd.notnull(x) else x)

    output_dir = 'data/output'
    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df_modelo_processo, index=False, header=True):
        ws.append(row)
    header_row = ws[1]
    for cell in header_row:
        cell.font = Font(name='Arial', bold=True, size=10)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
    file_path = os.path.join(output_dir, 'PROCESSOS.xlsx')
    wb.save(file_path)

# trasformacao_clientes()
# transformacao_processos()
